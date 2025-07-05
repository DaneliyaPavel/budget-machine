from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Iterable
from uuid import UUID

from openpyxl import load_workbook

from fastapi import (
    APIRouter,
    Depends,
    UploadFile,
    File,
    status,
    Query,
    Response,
    HTTPException,
)
from sqlalchemy.ext.asyncio import AsyncSession

from ... import crud, schemas, database
from ...services import ledger
from ...models import User
from ..utils import api_error
from .users import get_current_user

router = APIRouter(prefix="/transactions", tags=["Операции"])


@router.get("/", response_model=list[schemas.Transaction])
async def read_transactions(
    date_from: datetime | None = Query(None, description="Start date (posted_at)"),
    date_to: datetime | None = Query(None, description="End date (posted_at)"),
    category_id: str | None = Query(None, description="Category"),
    account_id: UUID | None = Query(None, description="Account"),
    limit: int = Query(50, description="Limit"),
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Вернуть список операций с указанными фильтрами."""
    return await crud.get_transactions(
        session,
        account_id or current_user.account_id,
        date_from=date_from,
        date_to=date_to,
        category_id=category_id,
        limit=limit,
    )


@router.post("/", response_model=schemas.Transaction)
async def create_transaction(
    tx: schemas.TransactionCreate,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Создать новую операцию."""
    if current_user.role == "readonly":
        raise api_error(status.HTTP_403_FORBIDDEN, "Forbidden", "FORBIDDEN")
    return await ledger.post_entry(
        session,
        tx,
        tx.postings,
        current_user.account_id,
        current_user.id,
    )


def _parse_rows(rows: Iterable[dict]) -> list[schemas.TransactionCreate]:
    parsed: list[schemas.TransactionCreate] = []
    for row in rows:
        try:
            posted_at = (
                datetime.fromisoformat(str(row.get("posted_at")))
                if row.get("posted_at")
                else None
            )
        except ValueError:
            posted_at = None
        parsed.append(
            schemas.TransactionCreate(
                payee=row.get("payee"),
                note=row.get("note"),
                category_id=str(row["category_id"]),
                posted_at=posted_at,
            )
        )
    return parsed


@router.post("/import", status_code=status.HTTP_201_CREATED)
async def import_transactions(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Импортировать операции из CSV или Excel."""
    if current_user.role == "readonly":
        raise api_error(status.HTTP_403_FORBIDDEN, "Forbidden", "FORBIDDEN")
    content = await file.read()
    created: list[schemas.TransactionCreate]

    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = (
            dict(zip(headers, (cell.value for cell in row)))
            for row in ws.iter_rows(min_row=2)
        )
        created = _parse_rows(rows)
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        created = _parse_rows(reader)

    await crud.create_transactions_bulk(
        session, created, current_user.account_id, current_user.id
    )
    return {"created": len(created)}


@router.get("/export")
async def export_transactions(
    date_from: datetime | None = Query(None, description="Start date (posted_at)"),
    date_to: datetime | None = Query(None, description="End date (posted_at)"),
    category_id: str | None = Query(None, description="Category"),
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Экспортировать операции в CSV."""
    rows = await crud.get_transactions(
        session,
        current_user.account_id,
        date_from=date_from,
        date_to=date_to,
        category_id=category_id,
    )
    fieldnames = [
        "id",
        "payee",
        "note",
        "category_id",
        "posted_at",
        "user_id",
    ]
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    for tx in rows:
        writer.writerow(
            {
                "id": tx.id,
                "payee": tx.payee or "",
                "note": tx.note or "",
                "category_id": tx.category_id,
                "posted_at": tx.posted_at.isoformat(),
                "user_id": tx.user_id,
            }
        )
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"},
    )


@router.get("/{tx_id}", response_model=schemas.Transaction)
async def read_transaction(
    tx_id: UUID,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Получить одну операцию."""
    tx = await crud.get_transaction(session, tx_id, current_user.account_id)
    if not tx:
        raise api_error(404, "Transaction not found", "TRANSACTION_NOT_FOUND")
    return tx


@router.patch("/{tx_id}", response_model=schemas.Transaction)
async def update_transaction(
    tx_id: UUID,
    data: schemas.TransactionUpdate,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Обновить операцию."""
    if current_user.role == "readonly":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={"detail": "Forbidden", "code": "FORBIDDEN"},
        )
    tx = await crud.update_transaction(session, tx_id, data, current_user.account_id)
    if not tx:
        raise api_error(404, "Transaction not found", "TRANSACTION_NOT_FOUND")
    return tx


@router.delete("/{tx_id}", status_code=204)
async def delete_transaction(
    tx_id: UUID,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Удалить операцию."""
    if current_user.role == "readonly":
        raise api_error(status.HTTP_403_FORBIDDEN, "Forbidden", "FORBIDDEN")
    await crud.delete_transaction(session, tx_id, current_user.account_id)
    return None
