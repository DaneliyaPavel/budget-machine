"""Маршруты для операций."""

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    UploadFile,
    File,
    status,
    Query,
    Response,
)
from sqlalchemy.ext.asyncio import AsyncSession
import csv
import io
from datetime import datetime
from typing import Iterable

from openpyxl import load_workbook
from .. import crud, schemas, database
from ..models import User
from .users import get_current_user

router = APIRouter(prefix="/операции", tags=["Операции"])


@router.get("/", response_model=list[schemas.Transaction])
async def read_transactions(
    start: datetime | None = Query(None, description="Начало периода"),
    end: datetime | None = Query(None, description="Конец периода"),
    category_id: str | None = Query(None, description="Категория"),
    limit: int | None = Query(None, description="Количество записей"),
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Получить список операций с фильтрами по дате и категории."""
    return await crud.get_transactions(
        session,
        current_user.account_id,
        start=start,
        end=end,
        category_id=category_id,
        limit=limit,
    )


@router.post("/", response_model=schemas.Transaction)
async def create_transaction(
    tx: schemas.TransactionCreate,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Создать новую операцию."""
    return await crud.create_transaction(
        session,
        tx,
        current_user.account_id,
        current_user.id,
    )


def _parse_rows(rows: Iterable[dict]) -> list[schemas.TransactionCreate]:
    parsed: list[schemas.TransactionCreate] = []
    for row in rows:
        try:
            created_at = (
                datetime.fromisoformat(str(row.get("created_at")))
                if row.get("created_at")
                else None
            )
        except ValueError:
            created_at = None
        parsed.append(
            schemas.TransactionCreate(
                amount=float(row["amount"]),
                currency=row.get("currency", "RUB"),
                description=row.get("description"),
                category_id=int(row["category_id"]),
                created_at=created_at,
            )
        )
    return parsed


@router.post("/импорт", status_code=status.HTTP_201_CREATED)
async def import_transactions(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Импортировать операции из CSV или Excel."""
    content = await file.read()
    created: list[schemas.TransactionCreate]

    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = (
            dict(zip(headers, (cell.value for cell in row)))
            for row in ws.iter_rows(min_row=2)
        )
        created = _parse_rows(rows)
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        created = _parse_rows(reader)

    await crud.create_transactions_bulk(
        session, created, current_user.account_id, current_user.id
    )
    return {"created": len(created)}


@router.get("/экспорт")
async def export_transactions(
    start: datetime | None = Query(None, description="Начало периода"),
    end: datetime | None = Query(None, description="Конец периода"),
    category_id: str | None = Query(None, description="Категория"),
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Выгрузить операции в формате CSV."""
    rows = await crud.get_transactions(
        session,
        current_user.account_id,
        start=start,
        end=end,
        category_id=category_id,
    )
    fieldnames = [
        "id",
        "amount",
        "currency",
        "amount_rub",
        "description",
        "category_id",
        "created_at",
        "account_id",
        "user_id",
    ]
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    for tx in rows:
        writer.writerow(
            {
                "id": tx.id,
                "amount": float(tx.amount),
                "currency": tx.currency,
                "amount_rub": float(tx.amount_rub),
                "description": tx.description,
                "category_id": tx.category_id,
                "created_at": tx.created_at.isoformat(),
                "account_id": tx.account_id,
                "user_id": tx.user_id,
            }
        )
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"},
    )


@router.get("/{tx_id}", response_model=schemas.Transaction)
async def read_transaction(
    tx_id: str,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Получить одну операцию."""
    tx = await crud.get_transaction(session, tx_id, current_user.account_id)
    if not tx:
        raise HTTPException(status_code=404, detail="Операция не найдена")
    return tx


@router.patch("/{tx_id}", response_model=schemas.Transaction)
async def update_transaction(
    tx_id: str,
    data: schemas.TransactionUpdate,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Изменить операцию."""
    tx = await crud.update_transaction(session, tx_id, data, current_user.account_id)
    if not tx:
        raise HTTPException(status_code=404, detail="Операция не найдена")
    return tx


@router.delete("/{tx_id}", status_code=204)
async def delete_transaction(
    tx_id: str,
    session: AsyncSession = Depends(database.get_session),
    current_user: User = Depends(get_current_user),
):
    """Удалить операцию."""
    await crud.delete_transaction(session, tx_id, current_user.account_id)
    return None
